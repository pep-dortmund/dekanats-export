from getpass import getpass
import os
import sys
import tempfile

import requests
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from datetime import datetime
from zoneinfo import ZoneInfo

from smtplib import SMTP_SSL
import logging
import os
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.utils import COMMASPACE, formatdate, formataddr

load_dotenv()

message = '''
Guten Tag Liebes Dekanat,

Anbei der aktuelle Stand der Anmeldungen für die 50-Jahr-Feier.

Mit freundlichen Grüßen
Maximilian Linhoff
'''


tz = ZoneInfo('Europe/Berlin')



URL_TEMPLATE = "https://registration.pep-dortmund.org/events/{event}/participants"


def get_participants():

    user = os.getenv("PEP_USER") or input("Username for registration.pep-dortmund.org: ")
    password = os.getenv("PEP_PASSWORD") or getpass("Password: ")
    event = os.getenv("PEP_EVENT", "physik50")

    r = requests.get(
        URL_TEMPLATE.format(event=event),
        auth=(user, password),
        headers={"Accept": "application/json"},
    )
    r.raise_for_status()

    return r.json()["participants"]


def save_excel(outputpath):
    '''
    {'email': 'hannes.kurtze@hs-anhalt.de',
     'freitag': {'begleitung_freitag': False, 'dinner': True, 'symp': True},
     'name': 'Hannes Kurtze',
     'samstag': {'BuB': True, 'bbq': False, 'begleitung_samstag': 0},
     'submit': True}
    '''

    participants = get_participants()
    if len(participants) == 0:
        print('Keine Anmeldungen bis jetzt', file=sys.stderr)
        sys.exit(1)

    participants = [
        {
            "Name": p["data"]["name"],
            "Anmeldng bestätigt": p["status_name"] == "confirmed",
            "Email": p["data"]["email"],
            "Festkolloquium": p["data"]["freitag"]["symp"],
            "Festessen": p["data"]["freitag"]["symp"],
            "Begleitung Festessen": int(p["data"]["freitag"]["symp"]),
            "Brötchen und Borussia": p["data"]["samstag"]["BuB"],
            "Alumnigrillen & Laborführungen": p["data"]["samstag"]["bbq"],
            "Anzahl Begleitpersonen": p["data"]["samstag"]["begleitung_samstag"],
        }
        for p in participants
    ]

    bold = Font("Cambria", bold=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Anmeldungen 50 Jahre Physik"
    ws['A1'] = f'Anmeldungen 50 Jahre Physik, Stand {datetime.now(tz).isoformat()}'
    ws['A1'].font = bold
    ws.merge_cells(f'A1:{get_column_letter(len(participants[0]))}1')


    col_widths = {}
    for col, colname in enumerate(participants[0].keys(), start=1):
        cell = ws.cell(row=2, column=col, value=colname)
        cell.font = bold
        col_widths[col] = 1.1 * len(colname)

    for row, participant in enumerate(participants, start=3):
        for col, value in enumerate(participant.values(), start=1):
            ws.cell(row=row, column=col, value=value)
            col_widths[col] = max(col_widths[col], len(str(value)))

    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    wb.save(outputpath)


def send_mail():
    msg = MIMEMultipart()
    from_addr = formataddr(("PeP et al. e.V.", "no-reply@pep-dortmund.org"))
    to_addrs = [
        "dekanat.physik@tu-dortmund.de",
        "maximilian.linhoff@tu-dortmund.de",
    ]
    msg['From'] = from_addr
    msg['Reply-To'] = formataddr(("Maximilian Linhoff", "maximilian.linhoff@tu-dortmund.de"))
    msg['To'] = COMMASPACE.join(to_addrs)
    msg['Date'] = formatdate(localtime=True)
    msg['Subject'] = "Update Anmeldungen 50-Jahr-Feier"
    msg.attach(MIMEText(message))

    name = "Anmeldungen50JahrePhysik.xlsx"
    with tempfile.NamedTemporaryFile(suffix=".xlsx") as tmp:
        save_excel(tmp.name)
        part = MIMEApplication(tmp.read(), name=name)
    part['Content-Disposition'] = f'attachment; filename="{name}"'
    msg.attach(part)


    server = os.getenv("PEP_MAIL_SERVER", "unimail.tu-dortmund.de")
    port = int(os.getenv("PEP_MAIL_PORT", 465))
    user = os.getenv("PEP_MAIL_USER") or input("Mail Username: ")
    password = os.getenv("PEP_MAIL_PASSWORD") or getpass("Mail Password: ")

    with SMTP_SSL(server, port) as smtp:
        logging.info('Connecting')
        smtp.login(user, password)
        logging.info('Login done. Sending message.')
        smtp.sendmail(from_addr, to_addrs=to_addrs, msg=msg.as_string())
        logging.info('Message sent')

if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    logging.info("Scheduling mail sending")
    send_mail()
